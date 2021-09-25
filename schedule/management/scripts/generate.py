import re
from dataclasses import dataclass
from typing import Any, Dict, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from schedule.models import FullTimeSchedule, Group, Schedule, Teacher


@dataclass
class Config:
    template_path: str
    from_week: int
    to_week: int
    print_item_name: bool
    print_item_type: bool
    print_places: bool
    print_groups: Optional[bool] = True
    print_teachers: Optional[bool] = True
    schedule_type: Optional[str] = Schedule.STUDY


def gen_groups_table(groups: Iterable[Group], config: Config) -> str:
    filename = "groups"
    wb = load_workbook(config.template_path)
    sh: Worksheet = wb["all"]

    for index, group in enumerate(groups):
        col = 2 * index + 4

        sh.cell(1, col, value=group.name)

        for week in range(config.to_week, config.from_week - 1, -1):
            item: FullTimeSchedule
            for item in FullTimeSchedule.objects.filter(
                group=group, week=week
            ).order_by("week", "date"):
                if not item.name:
                    continue

                date = item.date
                if not date:
                    raise ValueError(f"Item {item} has no date")

                item_type = item.item_type if config.print_item_type else ""
                item_name = item.name if config.print_item_name else ""
                place = item.place if config.print_places else ""
                teachers = (
                    ", ".join(str(teacher) for teacher in item.teachers.all())
                    if config.print_teachers
                    else ""
                )

                if len(item_name) > 57:
                    name_splitted = re.split(r"[ \-,.]", item_name)
                    name_uppercased = map(
                        lambda x: x[0].upper() if len(x) > 2 else x[0], name_splitted
                    )
                    item_name = "".join(name_uppercased)

                value = f"{item_name} {item_type}"
                if teachers:
                    value += f"\n{teachers}"

                try:
                    fill_items(
                        sh,
                        col,
                        content=value,
                        place=place,
                        week_day=item.week_day,
                        time=item.time,
                        week=week,
                    )
                except ValueError:
                    continue
    wb.save(f"{filename}.xlsx")
    return filename


def gen_teachers_table(
    teachers: Iterable[Teacher],
    config: Config,
) -> str:
    filename = "teachers"
    wb = load_workbook(config.template_path)
    sh: Worksheet = wb["all"]

    for index, teacher in enumerate(teachers):
        col = 2 * index + 4

        sh.cell(1, col, value=str(teacher))

        for week in range(config.to_week, config.from_week, -1):
            items: Dict[str, Dict[str, Any]] = {}
            item: FullTimeSchedule

            filter_cond = {
                "teachers__exact": teacher,
                "schedule_type": config.schedule_type,
            }
            if config.schedule_type == Schedule.STUDY:
                filter_cond["week"] = week  # type: ignore
            for item in FullTimeSchedule.objects.filter(**filter_cond):
                try:
                    existing = items[item.key()]
                except KeyError:
                    items[item.key()] = {
                        "date": item.date,
                        "week_day": item.week_day,
                        "time": item.time,
                        "groups": frozenset((str(item.group),)),
                        "item_type": item.item_type,
                        "name": item.name,
                        "place": item.place,
                    }
                else:
                    existing["groups"] |= {str(item.group)}

            for _, value in items.items():
                week_day = value["week_day"]
                groups = (
                    "\n" + ", ".join(sorted(value["groups"]))
                    if config.print_groups
                    else ""
                )
                item_type = value["item_type"] if config.print_item_type else ""
                item_name = value["name"] if config.print_item_name else ""
                place = value["place"] if config.print_places else ""
                if not item_name:
                    continue

                if len(item_name) > 57:
                    name_splitted = re.split(r"[ \-,.]", item_name)
                    name_uppercased = map(
                        lambda x: x[0].upper() if len(x) > 2 else x[0], name_splitted
                    )
                    item_name = "".join(name_uppercased)

                try:
                    fill_items(
                        sh,
                        col,
                        content=f"{item_name} {item_type}{groups}",
                        place=place,
                        week_day=week_day,
                        time=value["time"],
                        week=week,
                    )
                except ValueError:
                    continue

    wb.save(f"{filename}.xlsx")
    return filename


def fill_items(sheet: Worksheet, col: int, **kwargs) -> None:
    content = kwargs.get("content")
    place = kwargs.get("place")
    week_day = kwargs.get("week_day")
    time = kwargs.get("time")
    if not time:
        raise ValueError(f"Not found time at {content}")

    week = kwargs.get("week")
    if not week:
        raise ValueError(f"Not found week at {content}")

    week_offset = abs((week - 1) % 2)

    if week_day == "Пн":
        day_offset = 0 * 14
    elif week_day == "Вт":
        day_offset = 1 * 14
    elif week_day == "Ср":
        day_offset = 2 * 14
    elif week_day == "Чт":
        day_offset = 3 * 14
    elif week_day == "Пт":
        day_offset = 4 * 14
    elif week_day == "Сб":
        day_offset = 5 * 14
    else:
        raise ValueError(f"Unexpected day: {week_day}")

    if time == "09:00 – 10:30":
        row = 2
    elif time == "10:45 – 12:15":
        row = 4
    elif time == "13:00 – 14:30":
        row = 6
    elif time == "14:45 – 16:15":
        row = 8
    elif time == "16:30 – 18:00":
        row = 10
    elif time == "18:15 – 19:45":
        row = 12
    elif time == "20:00 – 21:30":
        row = 14
    else:
        raise ValueError(f"Unexpected time: {time}")
    row += day_offset + week_offset

    sheet.cell(row=row, column=col, value=content)
    sheet.cell(row=row, column=col + 1, value=place)
