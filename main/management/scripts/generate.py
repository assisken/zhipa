import re
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from main.models import Group, FullTimeSchedule, Teacher


def gen_groups_table(groups: Iterable[Group], from_week: int) -> str:
    filename = 'groups'
    wb = load_workbook('template.xlsx')
    sh: Worksheet = wb['all']

    for index, group in enumerate(groups):
        col = 2 * index + 4

        sh.cell(1, col, value=group.name)

        for week in range(17, from_week - 1, -1):
            for item in FullTimeSchedule.objects.filter(groups__exact=group, day__week=week).order_by('day__date'):
                day = item.day
                item_type = item.type
                name = item.name
                if not name:
                    continue
                place = '\n'.join(str(place) for place in item.places.all())
                teachers = ', '.join(str(teacher) for teacher in item.teachers.all())

                if len(name) > 57:
                    _name = re.split(r'[ \-,.]', name)
                    _name = map(lambda x: x[0].upper() if len(x) > 2 else x[0], _name)
                    name = ''.join(_name)

                value = f'{name} {item_type}'
                if teachers:
                    value += f'\n{teachers}'

                fill_items(sh, col,
                           content=value,
                           place=place,
                           day=day.week_day,
                           time_start=item.starts_at,
                           time_end=item.ends_at,
                           week=week)
    wb.save(f'{filename}.xlsx')
    return filename


def gen_teachers_table(teachers: Iterable[Teacher]) -> str:
    filename = 'teachers.xlsx'
    wb = load_workbook('template1.xlsx')
    sh: Worksheet = wb['all']

    for index, teacher in enumerate(teachers):
        col = 2 * index + 4

        sh.cell(1, col, value=str(teacher))

        for week in range(17, 0, -1):
            for item in FullTimeSchedule.objects.filter(teachers__exact=teacher, day__week=week):
                item: FullTimeSchedule
                day = item.day
                groups = ', '.join(group.name for group in item.groups.only('name'))
                item_type = item.item_type
                name = item.name
                if not name:
                    continue
                place = '\n'.join(str(place) for place in item.places.all())

                if len(name) > 57:
                    _name = re.split(r'[ \-,.]', name)
                    _name = map(lambda x: x[0].upper() if len(x) > 2 else x[0], _name)
                    name = ''.join(_name)

                fill_items(sh, col,
                           content=f'{name} {item_type}\n{groups}',
                           place=place,
                           day=day.day,
                           time_start=item.starts_at,
                           time_end=item.ends_at,
                           week=week)

    wb.save(filename)
    return filename


def fill_items(sheet: Worksheet, col: int, **kwargs) -> None:
    content = kwargs.get('content')
    place = kwargs.get('place')
    day = kwargs.get('day')
    time_start = kwargs.get('time_start')
    time_end = kwargs.get('time_end')
    time = '{} – {}'.format(time_start.strftime('%H:%M'), time_end.strftime('%H:%M'))
    week = kwargs.get('week')

    week_offset = abs((week - 1) % 2)

    if day == 'Пн':
        day_offset = 0 * 14
    elif day == 'Вт':
        day_offset = 1 * 14
    elif day == 'Ср':
        day_offset = 2 * 14
    elif day == 'Чт':
        day_offset = 3 * 14
    elif day == 'Пт':
        day_offset = 4 * 14
    elif day == 'Сб':
        day_offset = 5 * 14
    else:
        raise Exception(f'Unexpected day: {day}')

    if time == '09:00 – 10:30':
        row = 2
    elif time == '10:45 – 12:15':
        row = 4
    elif time == '13:00 – 14:30':
        row = 6
    elif time == '14:45 – 16:15':
        row = 8
    elif time == '16:30 – 18:00':
        row = 10
    elif time == '18:15 – 19:45':
        row = 12
    elif time == '20:00 – 21:30':
        row = 14
    else:
        raise Exception(f'Unexpected time: {time}')
    row += day_offset + week_offset

    sheet.cell(row=row, column=col, value=content)
    sheet.cell(row=row, column=col + 1, value=place)
